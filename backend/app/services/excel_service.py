from __future__ import annotations

from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
import re
from typing import Any

import pandas as pd
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_ILLEGAL_SHEET_CHARS = re.compile(r"[\[\]\*:/\\?]")
_ILLEGAL_XML_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_WHITESPACE = re.compile(r"\s+")

_HEADER_FILL = PatternFill("solid", fgColor="1F5A7A")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


@dataclass(slots=True)
class ParsedSheet:
    name: str
    headers: list[str]
    rows: list[dict[str, Any]]


def _sanitize_xml_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return _ILLEGAL_XML_CHARS.sub("", text)


def _sanitize_sheet_name(name: str, fallback: str = "Sheet") -> str:
    cleaned = _ILLEGAL_SHEET_CHARS.sub(" ", _sanitize_xml_text(name))
    cleaned = _WHITESPACE.sub(" ", cleaned).strip()
    return (cleaned or fallback)[:31]


def _unique_sheet_name(base_name: str, existing_names: set[str]) -> str:
    candidate = _sanitize_sheet_name(base_name)
    if candidate not in existing_names:
        existing_names.add(candidate)
        return candidate

    index = 2
    while True:
        suffix = f" {index}"
        trimmed = candidate[: 31 - len(suffix)]
        next_name = f"{trimmed}{suffix}".strip()
        if next_name not in existing_names:
            existing_names.add(next_name)
            return next_name
        index += 1


def _normalize_scalar(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, pd.Timedelta):
        return str(value)
    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            value = value.item()
        except Exception:
            pass
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return _sanitize_xml_text(value)
    return value


def _coerce_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = df.copy()
    cleaned.columns = [str(column).strip() if str(column).strip() else f"Column {index + 1}" for index, column in enumerate(cleaned.columns)]
    return cleaned


def _frame_to_rows(df: pd.DataFrame, limit: int | None = None) -> list[dict[str, Any]]:
    if limit is not None:
        df = df.head(limit)
    records: list[dict[str, Any]] = []
    for record in df.to_dict(orient="records"):
        records.append({key: _normalize_scalar(value) for key, value in record.items()})
    return records


def _dataframe_to_table(df: pd.DataFrame) -> list[list[Any]]:
    table: list[list[Any]] = [list(df.columns)]
    for row in df.to_dict(orient="records"):
        table.append([_normalize_scalar(row.get(column)) for column in df.columns])
    return table


def _load_excel(content: bytes) -> tuple[pd.ExcelFile, list[str]]:
    try:
        excel_file = pd.ExcelFile(BytesIO(content), engine="openpyxl")
    except Exception as exc:
        raise HTTPException(status_code=400, detail="The uploaded file is not a valid xlsx workbook.") from exc
    sheet_names = list(excel_file.sheet_names)
    if not sheet_names:
        raise HTTPException(status_code=400, detail="The workbook does not contain any sheets.")
    return excel_file, sheet_names


def _read_sheet(excel_file: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl", dtype=object)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' was not found.") from exc
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to read sheet '{sheet_name}'.") from exc

    if df.empty and len(df.columns) == 0:
        raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' is empty.")
    return _coerce_dataframe(df)


def inspect_workbook(content: bytes, sheet_name: str | None = None) -> dict[str, Any]:
    excel_file, sheet_names = _load_excel(content)
    selected_sheet_name = sheet_name or sheet_names[0]
    if selected_sheet_name not in sheet_names:
        raise HTTPException(status_code=400, detail=f"Sheet '{selected_sheet_name}' was not found.")

    df = _read_sheet(excel_file, selected_sheet_name)
    return {
        "sheet_names": sheet_names,
        "selected_sheet": selected_sheet_name,
        "columns": list(df.columns),
        "row_count": int(len(df.index)),
        "preview_rows": _frame_to_rows(df, limit=10),
    }


def parse_workbook(content: bytes, sheet_name: str | None = None) -> tuple[list[str], ParsedSheet]:
    excel_file, sheet_names = _load_excel(content)
    selected_sheet_name = sheet_name or sheet_names[0]
    if selected_sheet_name not in sheet_names:
        raise HTTPException(status_code=400, detail=f"Sheet '{selected_sheet_name}' was not found.")

    df = _read_sheet(excel_file, selected_sheet_name)
    return sheet_names, ParsedSheet(
        name=selected_sheet_name,
        headers=list(df.columns),
        rows=_frame_to_rows(df),
    )


def _ensure_key_columns(sheet: ParsedSheet, key_columns: list[str]) -> None:
    if not key_columns:
        raise HTTPException(status_code=400, detail="Select at least one key column.")
    missing = [column for column in key_columns if column not in sheet.headers]
    if missing:
        missing_text = ", ".join(missing)
        raise HTTPException(status_code=400, detail=f"Missing key columns: {missing_text}.")


def _group_rows(sheet: ParsedSheet, key_columns: list[str]) -> "OrderedDict[tuple[Any, ...], pd.DataFrame]":
    frame = pd.DataFrame(sheet.rows, columns=sheet.headers)
    if frame.empty:
        return OrderedDict()

    grouped: "OrderedDict[tuple[Any, ...], pd.DataFrame]" = OrderedDict()
    for key, group in frame.groupby(key_columns, dropna=False, sort=False):
        if not isinstance(key, tuple):
            key = (key,)
        grouped[key] = group.reset_index(drop=True)
    return grouped


def _build_summary_rows(grouped: "OrderedDict[tuple[Any, ...], pd.DataFrame]", key_columns: list[str]) -> list[dict[str, Any]]:
    summary_rows: list[dict[str, Any]] = []
    for index, (key, group) in enumerate(grouped.items(), start=1):
        label = " / ".join(
            f"{column}={_normalize_scalar(value)}" for column, value in zip(key_columns, key, strict=False)
        )
        summary_rows.append({"index": index, "label": label, "row_count": int(len(group.index))})
    return summary_rows


def _joined_value(values: pd.Series) -> Any:
    non_null = [value for value in values.tolist() if not pd.isna(value) and value != ""]
    if not non_null:
        return None
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in non_null):
        total = float(sum(float(value) for value in non_null))
        return int(total) if total.is_integer() else total

    seen: list[str] = []
    for value in non_null:
        text = str(_normalize_scalar(value))
        if text not in seen:
            seen.append(text)
    return " / ".join(seen)


def _merge_group_rows(sheet: ParsedSheet, key_columns: list[str], grouped: "OrderedDict[tuple[Any, ...], pd.DataFrame]") -> list[dict[str, Any]]:
    merged_rows: list[dict[str, Any]] = []
    for key, group in grouped.items():
        merged: dict[str, Any] = {}
        for column, value in zip(key_columns, key, strict=False):
            merged[column] = _normalize_scalar(value)
        merged["row_count"] = int(len(group.index))
        for column in sheet.headers:
            if column in key_columns:
                continue
            merged[column] = _joined_value(group[column])
        ordered: dict[str, Any] = {}
        for column in key_columns + ["row_count"] + [column for column in sheet.headers if column not in key_columns]:
            if column in merged:
                ordered[column] = merged[column]
        merged_rows.append(ordered)
    return merged_rows


def _style_header_row(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _write_frame(worksheet, rows: list[list[Any]]) -> None:
    for row in rows:
        worksheet.append([_normalize_scalar(value) for value in row])
    if worksheet.max_row > 0:
        _style_header_row(worksheet)
        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


def _build_workbook(workbook_sheets: list[tuple[str, list[list[Any]]]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    existing_names: set[str] = set()

    for sheet_name, rows in workbook_sheets:
        unique_name = _unique_sheet_name(sheet_name, existing_names)
        worksheet = workbook.create_sheet(unique_name)
        _write_frame(worksheet, rows)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def summarize_split(sheet: ParsedSheet, key_columns: list[str]) -> dict[str, Any]:
    _ensure_key_columns(sheet, key_columns)
    grouped = _group_rows(sheet, key_columns)
    summary_rows = _build_summary_rows(grouped, key_columns)
    first_group_rows = next(iter(grouped.values()), pd.DataFrame(columns=sheet.headers))
    output_sheet_names = ["summary"]
    for index, (key, group) in enumerate(grouped.items(), start=1):
        if len(key_columns) == 1:
            label = f"{key_columns[0]}-{_normalize_scalar(key[0])}"
        else:
            label = " / ".join(f"{column}-{_normalize_scalar(value)}" for column, value in zip(key_columns, key, strict=False))
        output_sheet_names.append(_sanitize_sheet_name(f"{index:02d}-{label}"))

    return {
        "mode": "split",
        "sheet_name": sheet.name,
        "key_columns": key_columns,
        "group_count": len(grouped),
        "total_rows": len(sheet.rows),
        "summary_rows": summary_rows,
        "preview_rows": _frame_to_rows(first_group_rows, limit=5),
        "output_sheet_names": output_sheet_names,
    }


def summarize_merge(sheet: ParsedSheet, key_columns: list[str]) -> dict[str, Any]:
    _ensure_key_columns(sheet, key_columns)
    grouped = _group_rows(sheet, key_columns)
    merged_rows = _merge_group_rows(sheet, key_columns, grouped)
    merged_frame = pd.DataFrame(merged_rows)
    return {
        "mode": "merge",
        "sheet_name": sheet.name,
        "key_columns": key_columns,
        "group_count": len(grouped),
        "total_rows": len(sheet.rows),
        "summary_rows": _build_summary_rows(grouped, key_columns),
        "preview_rows": _frame_to_rows(merged_frame, limit=5),
        "output_sheet_names": ["merged_result"],
    }


def build_output_workbook(sheet: ParsedSheet, key_columns: list[str], mode: str) -> tuple[bytes, dict[str, Any]]:
    _ensure_key_columns(sheet, key_columns)
    grouped = _group_rows(sheet, key_columns)
    if not grouped:
        raise HTTPException(status_code=400, detail="No rows matched the selected keys.")

    if mode == "split":
        summary_rows = _build_summary_rows(grouped, key_columns)
        workbook_sheets: list[tuple[str, list[list[Any]]]] = [
            ("summary", [["index", "label", "row_count"]] + [[row["index"], row["label"], row["row_count"]] for row in summary_rows])
        ]
        for index, (key, group) in enumerate(grouped.items(), start=1):
            if len(key_columns) == 1:
                label = f"{key_columns[0]}-{_normalize_scalar(key[0])}"
            else:
                label = " / ".join(f"{column}-{_normalize_scalar(value)}" for column, value in zip(key_columns, key, strict=False))
            sheet_name = _sanitize_sheet_name(f"{index:02d}-{label}")
            workbook_sheets.append((sheet_name, _dataframe_to_table(group)))

        summary = {
            "mode": "split",
            "sheet_name": sheet.name,
            "key_columns": key_columns,
            "group_count": len(grouped),
            "total_rows": len(sheet.rows),
            "summary_rows": summary_rows,
            "preview_rows": _frame_to_rows(next(iter(grouped.values())), limit=5),
            "output_sheet_names": [name for name, _ in workbook_sheets],
        }
        return _build_workbook(workbook_sheets), summary

    if mode == "merge":
        merged_rows = _merge_group_rows(sheet, key_columns, grouped)
        headers = list(merged_rows[0].keys()) if merged_rows else key_columns + ["row_count"]
        workbook_sheets = [("merged_result", [headers] + [[row.get(header) for header in headers] for row in merged_rows])]
        summary = {
            "mode": "merge",
            "sheet_name": sheet.name,
            "key_columns": key_columns,
            "group_count": len(grouped),
            "total_rows": len(sheet.rows),
            "summary_rows": _build_summary_rows(grouped, key_columns),
            "preview_rows": _frame_to_rows(pd.DataFrame(merged_rows), limit=5),
            "output_sheet_names": ["merged_result"],
        }
        return _build_workbook(workbook_sheets), summary

    raise HTTPException(status_code=400, detail="Unsupported processing mode.")


def inspect_output_name(sheet_name: str, mode: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M")
    suffix = "merge" if mode == "merge" else "split"
    return f"{_sanitize_sheet_name(sheet_name)}-{suffix}-{timestamp}.xlsx"
